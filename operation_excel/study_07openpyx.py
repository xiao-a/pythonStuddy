# 1.安装openpyxl
# pip install openpyxl

from tkinter import W
import openpyxl

# 新建excel
def new(): 
    # 创建
    from openpyxl import Workbook
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    # ws = wb.active

    # 方式一:插入到最后default (自定义sheet名称)
    # ws1 = wb.create_sheet("Mysheet")

    # 方式二:插入到最开始的位置（自定义sheet名称）(第一个参数：位置)
    ws2 = wb.create_sheet("Mysheet2", 0)

    wb.save("new01.xlsx")

# 打开excel

def opne():
    # 打开已有
    from openpyxl import load_workbook
    wb = load_workbook("new01.xlsx")
    # 选择已有表
    # sheet 名称可以作为key进行索引
    ws1 = wb.active
    ws3 = wb['Sheet1']
    ws4 = wb.get_sheet_by_name("Sheet1")

def show_sheet():
    # 查看表名
    from openpyxl import load_workbook
    wb = load_workbook("new01.xlsx")
    print(wb.sheetnames)
    for s in wb:
        print(s.title)

def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook("new01.xlsx")
    ws = wb.active
    v1 = ws['B3']
    v2 = ws.cell(row=4, column=3)
    print(v1.value, v2.value)

def get_many_value():
    from openpyxl import load_workbook
    wb = load_workbook("new01.xlsx")
    ws = wb.active
    # 通过切片
    cell_range = ws['A1':'C2']
    # 通过行或者列
    column = ws['C']
    column2 = ws['C':'D']
    rows1 = ws[10]
    rows2 = ws[10:20]

    print(cell_range)
    print("------------")
    print(column)
    print("------------")
    print(column2)

    # 通过指定范围（行-》行） 3行3列
    

    # 遍历所有方法
    print(tuple(ws.rows))
    print(tuple(ws.columns))

def get_num():
    from openpyxl import load_workbook
    wb = load_workbook("new01.xlsx")
    ws = wb.active    
    print(ws.max_row)
    print(ws.min_row)



if __name__ == "__main__":
    # new()
    # show_sheet()
    # get_one_value()
    get_many_value()

